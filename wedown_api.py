# -*- coding: utf-8 -*-
"""
公众号文章批量下载与术语表生成工具 - 带术语识别与自动翻译功能
支持：免费翻译引擎、DeepSeek API翻译
"""

import sys
import os

# 添加本地libs目录到Python路径
libs_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'libs')
if os.path.exists(libs_path) and libs_path not in sys.path:
    sys.path.insert(0, libs_path)

import re
import json
import hashlib
import requests
import time
from collections import Counter
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Dict, Tuple, Optional, Any
import warnings
warnings.filterwarnings('ignore')

# PyQt5 导入
try:
    from PyQt5.QtWidgets import *
    from PyQt5.QtCore import *
    from PyQt5.QtGui import QFont
except ImportError:
    print("错误：未安装PyQt5，请运行: pip install PyQt5")
    sys.exit(1)

# 基础库导入
try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    print("错误：未安装requests或beautifulsoup4，请运行: pip install requests beautifulsoup4")
    sys.exit(1)

# wxmp 可选
try:
    from wxmp import WxMPAPI
    WXMP_AVAILABLE = True
except ImportError:
    WXMP_AVAILABLE = False
    print("警告：未安装wxmp库，下载功能不可用。请运行: pip install wxmp")

# nltk 可选
try:
    import nltk
    from nltk.tokenize import word_tokenize
    from nltk.tag import pos_tag
    nltk.download('punkt_tab', quiet=True)
    nltk.download('averaged_perceptron_tagger_eng', quiet=True)
    NLP_AVAILABLE = True
except ImportError:
    NLP_AVAILABLE = False
    print("警告：未安装nltk库，频率统计模式下NLP功能不可用")

# openpyxl 可选
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("提示：未安装openpyxl库，Excel导出功能不可用。如需使用，请运行: pip install openpyxl")


# ==================== 免费翻译引擎 ====================
class FreeTranslator:
    """免费的翻译引擎封装，无需 API Key"""
    
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        # 可用翻译服务列表
        self.services = [
            self._translate_via_google,      # Google 翻译（免费，无需密钥）
            self._translate_via_deeplx,      # DeepLX（免费，推荐）
            self._translate_via_libretranslate  # LibreTranslate（免费）
        ]
    
    def _translate_via_google(self, text: str, target_lang: str = "zh-CN") -> Optional[str]:
        """使用 Google 翻译（免费，无需密钥）"""
        try:
            url = "https://translate.googleapis.com/translate_a/single"
            params = {
                "client": "gtx",
                "sl": "auto",
                "tl": target_lang,
                "dt": "t",
                "q": text
            }
            resp = self.session.get(url, params=params, timeout=8)
            if resp.status_code == 200:
                result = resp.json()
                # 解析 Google 翻译返回的 JSON 格式
                if result and len(result) > 0:
                    translation = ''.join([part[0] for part in result[0] if part[0]])
                    return translation.strip()
            return None
        except Exception:
            return None
    
    def _translate_via_deeplx(self, text: str, target_lang: str = "ZH") -> Optional[str]:
        """使用 DeepLX（免费的 DeepL 替代，建议本地部署或使用公共实例）"""
        # 公共实例列表（可能不稳定，建议自托管）
        public_instances = [
            "https://deeplx.owo.network",
            "https://deeplx.leez.dev"
        ]
        for base_url in public_instances:
            try:
                url = f"{base_url}/translate"
                payload = {
                    "text": text,
                    "target_lang": target_lang
                }
                resp = self.session.post(url, json=payload, timeout=5)
                if resp.status_code == 200:
                    data = resp.json()
                    if data.get("code") == 200 and data.get("data"):
                        return data["data"]
            except Exception:
                continue
        return None
    
    def _translate_via_libretranslate(self, text: str, target_lang: str = "zh") -> Optional[str]:
        """使用 LibreTranslate（免费开源）"""
        # 公共实例列表
        public_instances = [
            "https://translate.mentality.rip",
            "https://libretranslate.com"
        ]
        for base_url in public_instances:
            try:
                url = f"{base_url}/translate"
                payload = {
                    "q": text,
                    "source": "en",
                    "target": target_lang,
                    "format": "text"
                }
                resp = self.session.post(url, json=payload, timeout=5)
                if resp.status_code == 200:
                    data = resp.json()
                    if "translatedText" in data:
                        return data["translatedText"]
            except Exception:
                continue
        return None
    
    def translate(self, text: str) -> str:
        """翻译文本，自动尝试多个翻译服务"""
        if not text or not text.strip():
            return ""
        
        text = text.strip()
        
        # 尝试各个翻译服务
        for service in self.services:
            try:
                result = service(text)
                if result:
                    return result
            except Exception:
                continue
        
        return "翻译失败"


# ==================== DeepSeek API 翻译器 ====================
class DeepSeekTranslator:
    """DeepSeek API 翻译器，需要 API Key"""
    
    API_URL = "https://api.deepseek.com/chat/completions"
    
    def __init__(self, api_key: str = "", model: str = "deepseek-chat"):
        self.api_key = api_key
        self.model = model
        self.enabled = bool(api_key and api_key.strip())
    
    def translate_batch(self, terms: List[str]) -> Dict[str, str]:
        """批量翻译术语列表，返回 {term: translation} 字典"""
        if not self.enabled or not terms:
            return {}
        
        # 构建请求内容
        term_list = "\n".join([f"- {t}" for t in terms])
        prompt = f"""请将以下英文术语逐个翻译成中文，保持术语的准确性和专业性。请直接以JSON格式返回翻译结果，格式如下：
{{"term1": "翻译1", "term2": "翻译2", ...}}

需要翻译的术语：
{term_list}

注意：只返回JSON，不要有其他内容。"""
        
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": "你是一个专业的中英翻译专家，专注于学术和科技领域的术语翻译。"},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.3,
            "max_tokens": 2000
        }
        
        try:
            resp = requests.post(self.API_URL, headers=headers, json=payload, timeout=30)
            if resp.status_code == 200:
                data = resp.json()
                content = data["choices"][0]["message"]["content"]
                # 提取 JSON 内容
                content = content.strip()
                if content.startswith("```json"):
                    content = content[7:]
                if content.endswith("```"):
                    content = content[:-3]
                content = content.strip()
                result = json.loads(content)
                return result
            else:
                print(f"DeepSeek API 错误: {resp.status_code}")
                return {}
        except Exception as e:
            print(f"DeepSeek API 调用失败: {e}")
            return {}
    
    def translate_single(self, term: str) -> str:
        """翻译单个术语"""
        if not self.enabled:
            return "未配置API"
        
        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": self.model,
            "messages": [
                {"role": "system", "content": "你是一个专业的中英翻译专家。"},
                {"role": "user", "content": f"请将以下英文术语翻译成中文，只返回翻译结果，不要有其他内容：{term}"}
            ],
            "temperature": 0.3,
            "max_tokens": 100
        }
        
        try:
            resp = requests.post(self.API_URL, headers=headers, json=payload, timeout=15)
            if resp.status_code == 200:
                data = resp.json()
                return data["choices"][0]["message"]["content"].strip()
            else:
                return ""
        except Exception:
            return ""


# ==================== 术语处理器（带翻译功能） ====================
class TermProcessor:
    """术语表生成器（支持频率统计和顺序提取两种模式，支持自动翻译）"""
    
    DEFAULT_STOPWORDS = {
        'the', 'a', 'an', 'and', 'of', 'to', 'in', 'for', 'on', 'with',
        'by', 'is', 'are', 'was', 'were', 'be', 'been', 'being', 'have',
        'has', 'had', 'having', 'do', 'does', 'did', 'doing', 'but', 'or',
        'so', 'for', 'at', 'from', 'as', 'into', 'through', 'during',
        'this', 'that', 'these', 'those', 'it', 'they', 'we', 'you', 'he',
        'she', 'it', 'them', 'us', 'his', 'her', 'their', 'its', 'our',
        'your', 'my', 'all', 'any', 'both', 'each', 'few', 'more', 'most',
        'other', 'some', 'such', 'no', 'nor', 'not', 'only', 'own', 'same',
        'than', 'then', 'thence', 'there', 'these', 'they', 'this', 'those',
        'through', 'until', 'unto', 'when', 'where', 'which', 'while', 'will',
        'with', 'would', 'can', 'could', 'may', 'might', 'must', 'should',
    }
    
    def __init__(self, min_word_len=3, min_freq=1, max_terms=50):
        self.min_word_len = min_word_len
        self.min_freq = min_freq
        self.max_terms = max_terms
        self.stopwords = self.DEFAULT_STOPWORDS.copy()
        self.format_type = 0  # 0: 英文在前, 1: 中文在前
        self.translator_mode = "free"  # free 或 deepseek
        self.deepseek_translator = DeepSeekTranslator()
        self.free_translator = FreeTranslator()
        self.translation_cache = {}  # 缓存翻译结果
        self.enable_auto_translate = False  # 是否启用自动翻译
    
    def add_stopwords(self, words: List[str]):
        self.stopwords.update(words)
    
    def set_deepseek_key(self, api_key: str):
        """设置 DeepSeek API Key"""
        self.deepseek_translator = DeepSeekTranslator(api_key)
    
    def _clean_text(self, text: str) -> str:
        text = re.sub(r'[^\w\s\'\']', ' ', text)
        text = re.sub(r'\s+', ' ', text)
        return text.strip()
    
    def _is_valid_term(self, word: str) -> bool:
        if len(word) < self.min_word_len:
            return False
        if word.lower() in self.stopwords:
            return False
        if word.isdigit():
            return False
        if re.match(r'^[\d\W]+$', word):
            return False
        return True
    
    def _translate_term(self, term: str) -> str:
        """翻译单个术语，使用缓存"""
        if not self.enable_auto_translate:
            return ""
        
        term_lower = term.lower()
        if term_lower in self.translation_cache:
            return self.translation_cache[term_lower]
        
        # 根据选择的模式进行翻译
        if self.translator_mode == "deepseek" and self.deepseek_translator.enabled:
            result = self.deepseek_translator.translate_single(term)
        else:
            result = self.free_translator.translate(term)
        
        if result:
            self.translation_cache[term_lower] = result
        return result or ""
    
    def _translate_terms_batch(self, terms: List[str]) -> Dict[str, str]:
        """批量翻译术语"""
        if not self.enable_auto_translate:
            return {}
        
        # 筛选未翻译的术语
        untranslated = [t for t in terms if t.lower() not in self.translation_cache]
        
        if self.translator_mode == "deepseek" and self.deepseek_translator.enabled and untranslated:
            # 使用 DeepSeek 批量翻译
            translations = self.deepseek_translator.translate_batch(untranslated)
            for term, trans in translations.items():
                if trans:
                    self.translation_cache[term.lower()] = trans
        elif untranslated:
            # 使用免费翻译引擎逐个翻译
            for term in untranslated:
                trans = self.free_translator.translate(term)
                if trans:
                    self.translation_cache[term.lower()] = trans
                time.sleep(0.1)  # 避免请求过快
        
        # 返回所有术语的翻译结果
        result = {}
        for t in terms:
            if t.lower() in self.translation_cache:
                result[t] = self.translation_cache[t.lower()]
            elif self.enable_auto_translate:
                result[t] = "翻译中..."
            else:
                result[t] = ""
        return result
    
    # ---------- 频率统计模式 ----------
    def extract_basic(self, text: str) -> List[Tuple[str, str, int]]:
        cleaned = self._clean_text(text)
        words = re.findall(r"[A-Za-z]+(?:'[A-Za-z]+)?", cleaned.lower())
        filtered_words = [w for w in words if self._is_valid_term(w)]
        counter = Counter(filtered_words)
        terms_list = [(term, "", count) for term, count in counter.most_common(self.max_terms) if count >= self.min_freq]
        
        # 自动翻译
        if self.enable_auto_translate and terms_list:
            term_strings = [t[0] for t in terms_list]
            translations = self._translate_terms_batch(term_strings)
            terms_list = [(t[0], translations.get(t[0], ""), t[2]) for t in terms_list]
        
        return terms_list
    
    def extract_nlp(self, text: str) -> List[Tuple[str, str, int]]:
        if not NLP_AVAILABLE:
            return self.extract_basic(text)
        cleaned = self._clean_text(text)
        words = word_tokenize(cleaned.lower())
        tagged = pos_tag(words)
        valid_tags = ['NN', 'NNS', 'NNP', 'NNPS', 'JJ', 'JJR', 'JJS']
        filtered_terms = []
        for word, tag in tagged:
            if any(tag.startswith(t) for t in valid_tags) and self._is_valid_term(word):
                filtered_terms.append(word)
        counter = Counter(filtered_terms)
        terms_list = [(term, "", count) for term, count in counter.most_common(self.max_terms) if count >= self.min_freq]
        
        # 自动翻译
        if self.enable_auto_translate and terms_list:
            term_strings = [t[0] for t in terms_list]
            translations = self._translate_terms_batch(term_strings)
            terms_list = [(t[0], translations.get(t[0], ""), t[2]) for t in terms_list]
        
        return terms_list
    
    # ---------- 顺序提取模式 ----------
    def _split_chinese_english(self, text: str) -> Tuple[str, str]:
        first_cn = re.search(r'[\u4e00-\u9fff]', text)
        first_en = re.search(r'[A-Za-z]', text)
        if not first_cn and not first_en:
            return ("", "")
        if not first_cn:
            return (text.strip(), "")
        if not first_en:
            return ("", text.strip())
        if first_cn.start() < first_en.start():
            cn_part = re.match(r'^[\u4e00-\u9fff]+', text)
            if cn_part:
                cn_text = cn_part.group()
                en_text = text[cn_part.end():].strip()
                return (en_text, cn_text)
        else:
            en_part = re.match(r'^[A-Za-z\s\'-]+', text)
            if en_part:
                en_text = en_part.group().strip()
                cn_text = text[en_part.end():].strip()
                return (en_text, cn_text)
        return ("", "")
    
    def extract_ordered_pairs(self, text: str) -> List[Tuple[str, str]]:
        pattern = r'(\d+)\.\s*([^\d]+?)(?=\d+\.|$)'
        matches = re.findall(pattern, text, re.DOTALL)
        pairs = []
        for num_str, content in matches:
            content = content.strip()
            if not content:
                continue
            has_cn = bool(re.search(r'[\u4e00-\u9fff]', content))
            has_en = bool(re.search(r'[A-Za-z]', content))
            if has_cn and has_en:
                en, cn = self._split_chinese_english(content)
                if en and cn:
                    pairs.append((en, cn))
                else:
                    pass
        seen = set()
        unique = []
        for en, cn in pairs:
            key = (en.lower(), cn)
            if key not in seen:
                seen.add(key)
                unique.append((en, cn))
        
        # 如果开启了自动翻译且有些条目没有中文，尝试补充翻译
        if self.enable_auto_translate:
            new_pairs = []
            for en, cn in unique:
                if not cn or cn.strip() == "":
                    translation = self._translate_term(en)
                    if translation:
                        new_pairs.append((en, translation))
                    else:
                        new_pairs.append((en, cn))
                else:
                    new_pairs.append((en, cn))
            return new_pairs
        
        return unique
    
    # ---------- 翻译功能 ----------
    def translate_terms(self, terms: List[Tuple[str, str, int]]) -> List[Tuple[str, str, int]]:
        """为频率统计模式的术语添加翻译"""
        if not self.enable_auto_translate or not terms:
            return terms
        
        term_strings = [t[0] for t in terms]
        translations = self._translate_terms_batch(term_strings)
        return [(t[0], translations.get(t[0], t[1]), t[2]) for t in terms]
    
    # ---------- 导出方法 ----------
    def generate_markdown_table(self, terms, title: str = "术语表", is_ordered: bool = False):
        if is_ordered:
            lines = [f"# {title}\n", "| 序号 | 英文 | 中文 |", "|------|------|------|"]
            for idx, (en, cn) in enumerate(terms, 1):
                lines.append(f"| {idx} | {en} | {cn} |")
        else:
            has_cn = any(cn for _, cn, _ in terms)
            if has_cn:
                lines = [f"# {title}\n", "| 序号 | 英文 | 中文 | 出现次数 |", "|------|------|------|----------|"]
                for idx, (en, cn, count) in enumerate(terms, 1):
                    lines.append(f"| {idx} | {en} | {cn} | {count} |")
            else:
                lines = [f"# {title}\n", "| 序号 | 术语 | 出现次数 |", "|------|------|----------|"]
                for idx, (en, _, count) in enumerate(terms, 1):
                    lines.append(f"| {idx} | {en} | {count} |")
        return "\n".join(lines)
    
    def generate_json(self, terms, metadata: Dict = None, is_ordered: bool = False):
        data = {"generated_at": datetime.now().isoformat(), "total_terms": len(terms), "terms": []}
        if is_ordered:
            for en, cn in terms:
                data["terms"].append({"en": en, "cn": cn})
        else:
            for en, cn, count in terms:
                item = {"en": en, "frequency": count}
                if cn:
                    item["cn"] = cn
                data["terms"].append(item)
        if metadata:
            data["metadata"] = metadata
        return json.dumps(data, indent=2, ensure_ascii=False)
    
    def generate_csv(self, terms, is_ordered: bool = False):
        if is_ordered:
            lines = ["序号,英文,中文"]
            for idx, (en, cn) in enumerate(terms, 1):
                en_escaped = f'"{en}"' if ',' in en else en
                cn_escaped = f'"{cn}"' if ',' in cn else cn
                lines.append(f"{idx},{en_escaped},{cn_escaped}")
        else:
            has_cn = any(cn for _, cn, _ in terms)
            if has_cn:
                lines = ["英文,中文,出现次数"]
                for en, cn, count in terms:
                    en_escaped = f'"{en}"' if ',' in en else en
                    cn_escaped = f'"{cn}"' if ',' in cn else cn
                    lines.append(f"{en_escaped},{cn_escaped},{count}")
            else:
                lines = ["术语,出现次数"]
                for en, _, count in terms:
                    en_escaped = f'"{en}"' if ',' in en else en
                    lines.append(f"{en_escaped},{count}")
        return "\n".join(lines)
    
    def generate_excel(self, terms, file_path: str, title: str = "术语表", is_ordered: bool = False):
        if not EXCEL_AVAILABLE:
            return False
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "术语表"
            if is_ordered:
                headers = ["序号", "英文", "中文"]
                ws.column_dimensions['A'].width = 8
                ws.column_dimensions['B'].width = 30
                ws.column_dimensions['C'].width = 20
            else:
                has_cn = any(cn for _, cn, _ in terms)
                if has_cn:
                    headers = ["序号", "英文", "中文", "出现次数"]
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 30
                    ws.column_dimensions['C'].width = 20
                    ws.column_dimensions['D'].width = 12
                else:
                    headers = ["序号", "术语", "出现次数"]
                    ws.column_dimensions['A'].width = 8
                    ws.column_dimensions['B'].width = 30
                    ws.column_dimensions['C'].width = 12
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="0078D7", end_color="0078D7", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            cell_alignment = Alignment(vertical="center")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            if is_ordered:
                for idx, (en, cn) in enumerate(terms, 1):
                    row = idx + 1
                    ws.cell(row=row, column=1, value=idx).border = thin_border
                    ws.cell(row=row, column=2, value=en).border = thin_border
                    ws.cell(row=row, column=3, value=cn).border = thin_border
            else:
                has_cn = any(cn for _, cn, _ in terms)
                for idx, (en, cn, count) in enumerate(terms, 1):
                    row = idx + 1
                    ws.cell(row=row, column=1, value=idx).border = thin_border
                    ws.cell(row=row, column=2, value=en).border = thin_border
                    if has_cn:
                        ws.cell(row=row, column=3, value=cn).border = thin_border
                        ws.cell(row=row, column=4, value=count).border = thin_border
                    else:
                        ws.cell(row=row, column=3, value=count).border = thin_border
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"Excel生成失败: {e}")
            return False


# ==================== 下载管理器 ====================
class DownloadManager(QObject):
    progress_signal = pyqtSignal(int, int)
    log_signal = pyqtSignal(str)
    article_downloaded = pyqtSignal(dict)
    download_finished = pyqtSignal()
    
    def __init__(self, cookies: Dict[str, str] = None):
        super().__init__()
        self.cookies = cookies
        self.api = None
        self._running = False
    
    def init_api(self, cookies_file: str = None):
        if not WXMP_AVAILABLE:
            self.log_signal.emit("错误：未安装wxmp库，无法执行下载")
            return False
        try:
            if cookies_file and os.path.exists(cookies_file):
                with open(cookies_file, 'r') as f:
                    cookies_data = json.load(f)
                self.api = WxMPAPI(cookies_data)
            elif self.cookies:
                self.api = WxMPAPI(self.cookies)
            else:
                self.log_signal.emit("错误：缺少cookies配置")
                return False
            self.log_signal.emit("API初始化成功")
            return True
        except Exception as e:
            self.log_signal.emit(f"API初始化失败: {str(e)}")
            return False
    
    def search_account(self, keyword: str) -> List[Dict]:
        if not self.api:
            self.log_signal.emit("请先初始化API")
            return []
        try:
            response = self.api.search_fakeid(keyword)
            results = []
            for item in response.list:
                results.append({
                    "name": item.nickname,
                    "fakeid": item.fakeid,
                    "alias": getattr(item, 'alias', '')
                })
            self.log_signal.emit(f"搜索到 {len(results)} 个公众号")
            return results
        except Exception as e:
            self.log_signal.emit(f"搜索失败: {str(e)}")
            return []
    
    def download_article(self, url: str, save_dir: str) -> Tuple[bool, str, str]:
        try:
            response = requests.get(url, headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            })
            response.encoding = 'utf-8'
            soup = BeautifulSoup(response.text, 'html.parser')
            title_elem = soup.find('h1', class_='rich_media_title')
            title = title_elem.get_text().strip() if title_elem else "无标题"
            content_elem = soup.find('div', class_='rich_media_content')
            text = content_elem.get_text().strip() if content_elem else ""
            if not text:
                self.log_signal.emit(f"无法提取文章内容: {url}")
                return False, "", title
            os.makedirs(save_dir, exist_ok=True)
            safe_title = re.sub(r'[\\/*?:"<>|]', '_', title)
            filepath = os.path.join(save_dir, f"{safe_title}.md")
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(f"# {title}\n\n")
                f.write(f"原文链接：{url}\n\n")
                f.write(text)
            self.log_signal.emit(f"已下载: {title}")
            return True, filepath, title
        except Exception as e:
            self.log_signal.emit(f"下载失败 {url}: {str(e)}")
            return False, "", ""
    
    def batch_download(self, urls: List[str], save_dir: str, max_workers: int = 3):
        self._running = True
        total = len(urls)
        results = []
        try:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {executor.submit(self.download_article, url, save_dir): url for url in urls}
                for idx, future in enumerate(as_completed(futures), 1):
                    if not self._running:
                        break
                    try:
                        success, path, title = future.result()
                    except Exception as e:
                        self.log_signal.emit(f"下载异常: {str(e)}")
                        success, path, title = False, "", ""
                    results.append({"url": futures[future], "success": success, "path": path, "title": title})
                    self.progress_signal.emit(idx, total)
                    self.article_downloaded.emit({"title": title, "path": path, "success": success})
        except Exception as e:
            self.log_signal.emit(f"批量下载异常: {str(e)}")
        finally:
            self._running = False
            self.log_signal.emit(f"批量下载完成，成功 {sum(1 for r in results if r['success'])}/{total}")
            self.download_finished.emit()
        return results
    
    def stop(self):
        self._running = False
        self.log_signal.emit("正在停止下载...")


class DownloadWorker(QThread):
    finished = pyqtSignal()
    def __init__(self, downloader: DownloadManager, urls: List[str], save_dir: str, max_workers: int):
        super().__init__()
        self.downloader = downloader
        self.urls = urls
        self.save_dir = save_dir
        self.max_workers = max_workers
    def run(self):
        self.downloader.batch_download(self.urls, self.save_dir, self.max_workers)
        self.finished.emit()


# ==================== 主窗口 ====================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("公众号文章批量下载与术语表生成工具 - 带术语翻译")
        self.resize(1100, 700)
        self.setMinimumSize(800, 600)
        self.setStyleSheet("""
            QMainWindow { background-color: #f5f5f5; }
            QGroupBox { font-weight: bold; border: 1px solid #ccc; border-radius: 5px; margin-top: 10px; padding-top: 10px; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px 0 5px; }
            QLineEdit, QTextEdit, QListWidget, QComboBox { border: 1px solid #ccc; border-radius: 3px; padding: 3px; }
            QPushButton { background-color: #0078d7; color: white; border: none; border-radius: 3px; padding: 5px 15px; }
            QPushButton:hover { background-color: #106ebe; }
            QPushButton:disabled { background-color: #cccccc; }
            QTabWidget::pane { border: 1px solid #ccc; border-radius: 3px; }
            QTabBar::tab { background-color: #e0e0e0; padding: 5px 15px; margin-right: 2px; }
            QTabBar::tab:selected { background-color: #0078d7; color: white; }
        """)
        
        self.downloader = DownloadManager()
        self.term_processor = TermProcessor()
        self.download_results = []
        self.init_ui()
        self.setup_signals()
        self.log("[系统] 工具已启动")
        if not WXMP_AVAILABLE:
            self.log("[警告] wxmp库未安装，下载功能不可用。请运行: pip install wxmp")
        if not NLP_AVAILABLE:
            self.log("[警告] nltk库未安装，频率统计模式下NLP功能不可用")
        if not EXCEL_AVAILABLE:
            self.log("[提示] openpyxl库未安装，Excel导出功能不可用。请运行: pip install openpyxl")

    def init_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        self.tabs = QTabWidget()
        self.tabs.addTab(self.create_download_tab(), "📥 文章下载")
        self.tabs.addTab(self.create_term_tab(), "📊 术语表生成")
        self.tabs.addTab(self.create_settings_tab(), "⚙️ 设置")
        layout.addWidget(self.tabs)

    def create_download_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        url_group = QGroupBox("文章链接输入")
        url_layout = QVBoxLayout(url_group)
        self.url_input = QTextEdit()
        self.url_input.setPlaceholderText("请输入文章URL，每行一个...")
        self.url_input.setMaximumHeight(150)
        url_layout.addWidget(self.url_input)
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("包含关键词:"))
        self.include_keyword = QLineEdit()
        self.include_keyword.setPlaceholderText("多个关键词用逗号分隔")
        filter_layout.addWidget(self.include_keyword)
        filter_layout.addWidget(QLabel("排除关键词:"))
        self.exclude_keyword = QLineEdit()
        self.exclude_keyword.setPlaceholderText("多个关键词用逗号分隔")
        filter_layout.addWidget(self.exclude_keyword)
        filter_layout.addStretch()
        url_layout.addLayout(filter_layout)
        layout.addWidget(url_group)
        
        settings_group = QGroupBox("下载设置")
        settings_layout = QGridLayout(settings_group)
        settings_layout.addWidget(QLabel("保存格式:"), 0, 0)
        self.format_combo = QComboBox()
        self.format_combo.addItems(["Markdown (.md)", "HTML (.html)", "TXT (.txt)", "JSON (.json)", "所有格式"])
        settings_layout.addWidget(self.format_combo, 0, 1)
        settings_layout.addWidget(QLabel("保存路径:"), 0, 2)
        self.save_path = QLineEdit()
        self.save_path.setText("./downloads")
        settings_layout.addWidget(self.save_path, 0, 3)
        self.browse_btn = QPushButton("浏览")
        self.browse_btn.clicked.connect(self.browse_save_path)
        settings_layout.addWidget(self.browse_btn, 0, 4)
        settings_layout.addWidget(QLabel("并发下载数:"), 1, 0)
        self.max_workers = QSpinBox()
        self.max_workers.setRange(1, 5)
        self.max_workers.setValue(3)
        settings_layout.addWidget(self.max_workers, 1, 1)
        settings_layout.setColumnStretch(3, 2)
        layout.addWidget(settings_group)
        
        progress_group = QGroupBox("下载进度")
        progress_layout = QVBoxLayout(progress_group)
        self.progress_bar = QProgressBar()
        progress_layout.addWidget(self.progress_bar)
        self.status_label = QLabel("就绪")
        progress_layout.addWidget(self.status_label)
        layout.addWidget(progress_group)
        
        btn_layout = QHBoxLayout()
        self.download_btn = QPushButton("▶ 开始下载")
        self.download_btn.clicked.connect(self.start_download)
        self.stop_btn = QPushButton("⏹ 停止下载")
        self.stop_btn.clicked.connect(self.stop_download)
        self.stop_btn.setEnabled(False)
        btn_layout.addWidget(self.download_btn)
        btn_layout.addWidget(self.stop_btn)
        btn_layout.addStretch()
        self.import_btn = QPushButton("📁 导入已下载文章到术语表")
        self.import_btn.clicked.connect(self.import_downloaded_to_term)
        btn_layout.addWidget(self.import_btn)
        layout.addLayout(btn_layout)
        return tab

    def create_term_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # 输入区域
        input_group = QGroupBox("文章输入")
        input_layout = QVBoxLayout(input_group)
        self.term_input = QTextEdit()
        self.term_input.setPlaceholderText("请粘贴文章内容，或点击下方的'从下载的文件导入'按钮...")
        self.term_input.setMaximumHeight(200)
        input_layout.addWidget(self.term_input)
        
        load_layout = QHBoxLayout()
        self.load_articles_btn = QPushButton("📂 从下载目录加载已下载文章")
        self.load_articles_btn.clicked.connect(self.load_downloaded_articles)
        load_layout.addWidget(self.load_articles_btn)
        self.process_selected_btn = QPushButton("⚙️ 处理选中文章")
        self.process_selected_btn.clicked.connect(self.process_selected_articles)
        self.process_selected_btn.setEnabled(False)
        load_layout.addWidget(self.process_selected_btn)
        load_layout.addStretch()
        input_layout.addLayout(load_layout)
        layout.addWidget(input_group)
        
        # 文章列表
        articles_group = QGroupBox("已下载文章列表（可多选）")
        articles_layout = QVBoxLayout(articles_group)
        self.articles_list = QListWidget()
        self.articles_list.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.articles_list.itemDoubleClicked.connect(self.process_single_article)
        articles_layout.addWidget(self.articles_list)
        layout.addWidget(articles_group)
        
        # 模式选择
        mode_group = QGroupBox("术语提取模式")
        mode_layout = QHBoxLayout(mode_group)
        self.mode_freq = QRadioButton("频率统计模式（按词频排序）")
        self.mode_ordered = QRadioButton("顺序提取模式（按文章顺序编号）")
        self.mode_ordered.setChecked(True)
        mode_layout.addWidget(self.mode_freq)
        mode_layout.addWidget(self.mode_ordered)
        layout.addWidget(mode_group)
        
        # 格式选择
        format_group = QGroupBox("顺序提取模式格式")
        format_layout = QHBoxLayout(format_group)
        self.format_label = QLabel("条目格式:")
        self.format_combo_order = QComboBox()
        self.format_combo_order.addItems(["英文 中文 (如 'aristocracy 贵族')", "中文 英文 (如 '贵族 aristocracy')"])
        format_layout.addWidget(self.format_label)
        format_layout.addWidget(self.format_combo_order)
        format_layout.addStretch()
        layout.addWidget(format_group)
        
        # 翻译设置区域（新增）
        translate_group = QGroupBox("🔤 术语翻译设置")
        translate_layout = QGridLayout(translate_group)
        
        self.enable_translate_check = QCheckBox("启用术语自动翻译")
        self.enable_translate_check.setChecked(False)
        translate_layout.addWidget(self.enable_translate_check, 0, 0, 1, 2)
        
        translate_layout.addWidget(QLabel("翻译引擎:"), 1, 0)
        self.translate_engine = QComboBox()
        self.translate_engine.addItems(["免费翻译引擎（零配置，无需API密钥）", "DeepSeek API（需要API密钥，翻译质量更高）"])
        translate_layout.addWidget(self.translate_engine, 1, 1, 1, 2)
        
        translate_layout.addWidget(QLabel("DeepSeek API Key:"), 2, 0)
        self.deepseek_api_key_input = QLineEdit()
        self.deepseek_api_key_input.setPlaceholderText("输入你的 DeepSeek API Key")
        translate_layout.addWidget(self.deepseek_api_key_input, 2, 1, 1, 2)
        
        self.test_api_btn = QPushButton("测试API连接")
        self.test_api_btn.clicked.connect(self.test_deepseek_api)
        translate_layout.addWidget(self.test_api_btn, 3, 1)
        
        translate_info = QLabel("💡 免费翻译引擎无需配置即可使用，适合日常快速翻译。DeepSeek API 翻译质量更高，可批量翻译。")
        translate_info.setWordWrap(True)
        translate_info.setStyleSheet("color: #666; font-size: 11px;")
        translate_layout.addWidget(translate_info, 4, 0, 1, 3)
        
        layout.addWidget(translate_group)
        
        # 参数设置
        settings_group = QGroupBox("提取参数")
        settings_layout = QGridLayout(settings_group)
        
        # 第一列
        settings_layout.addWidget(QLabel("提取模式:"), 0, 0)
        self.extract_mode = QComboBox()
        self.extract_mode.addItems(["基础模式（频率统计）", "NLP模式（名词/形容词优先）"])
        settings_layout.addWidget(self.extract_mode, 0, 1)
        
        settings_layout.addWidget(QLabel("最小词长度:"), 1, 0)
        self.min_len = QSpinBox()
        self.min_len.setRange(2, 20)
        self.min_len.setValue(3)
        settings_layout.addWidget(self.min_len, 1, 1)
        
        settings_layout.addWidget(QLabel("最小出现次数:"), 2, 0)
        self.min_freq = QSpinBox()
        self.min_freq.setRange(1, 100)
        self.min_freq.setValue(2)
        settings_layout.addWidget(self.min_freq, 2, 1)
        
        # 第二列
        settings_layout.addWidget(QLabel("最大术语数量:"), 0, 2)
        self.max_terms = QSpinBox()
        self.max_terms.setRange(10, 500)
        self.max_terms.setValue(50)
        settings_layout.addWidget(self.max_terms, 0, 3)
        
        settings_layout.addWidget(QLabel("输出格式:"), 1, 2)
        self.output_format = QComboBox()
        self.output_format.addItems(["Markdown表格", "JSON", "CSV", "Excel (.xlsx)"])
        settings_layout.addWidget(self.output_format, 1, 3)
        
        settings_layout.addWidget(QLabel("自定义停用词:"), 2, 2)
        self.stopwords_input = QLineEdit()
        self.stopwords_input.setPlaceholderText("用逗号分隔，例如: test,example")
        settings_layout.addWidget(self.stopwords_input, 2, 3)
        layout.addWidget(settings_group)
        
        # 操作按钮
        btn_layout = QHBoxLayout()
        self.process_btn = QPushButton(" 提取术语")
        self.process_btn.clicked.connect(self.process_current_text)
        self.translate_btn = QPushButton("🌐 识别并翻译术语")
        self.translate_btn.clicked.connect(self.translate_current_terms)
        self.ai_extract_btn = QPushButton("🤖 AI术语提取")
        self.ai_extract_btn.clicked.connect(self.ai_extract_terms)
        self.ai_extract_btn.setStyleSheet("background-color: #28a745;")
        self.save_terms_btn = QPushButton("💾 保存术语表")
        self.save_terms_btn.clicked.connect(self.save_terms)
        btn_layout.addWidget(self.process_btn)
        btn_layout.addWidget(self.translate_btn)
        btn_layout.addWidget(self.ai_extract_btn)
        btn_layout.addWidget(self.save_terms_btn)
        btn_layout.addStretch()
        layout.addLayout(btn_layout)
        
        # 预览区域
        result_group = QGroupBox("术语表预览")
        result_layout = QVBoxLayout(result_group)
        self.result_preview = QTextEdit()
        self.result_preview.setReadOnly(True)
        self.result_preview.setFont(QFont("Consolas", 10))
        result_layout.addWidget(self.result_preview)
        layout.addWidget(result_group)
        return tab

    def create_settings_tab(self) -> QWidget:
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # 翻译服务说明
        translate_info_group = QGroupBox("🌐 翻译服务说明")
        translate_info_layout = QVBoxLayout(translate_info_group)
        info_text = QTextEdit()
        info_text.setReadOnly(True)
        info_text.setMaximumHeight(150)
        info_text.setPlainText("""【免费翻译引擎】
• 无需任何配置，开箱即用
• 免费无限调用
• 翻译质量中等，适合日常使用
• 技术支持：Google翻译 / DeepLX / LibreTranslate

【DeepSeek API】
• 需要 API Key，可访问 platform.deepseek.com 注册获取
• 新用户注册可领取免费额度
• 翻译质量高，支持上下文理解
• 适合专业术语、学术文献翻译""")
        translate_info_layout.addWidget(info_text)
        layout.addWidget(translate_info_group)
        
        # 更新连接状态区
        update_info_group = QGroupBox("📢 更新说明")
        update_info_layout = QVBoxLayout(update_info_group)
        update_text = QTextEdit()
        update_text.setReadOnly(True)
        update_text.setMaximumHeight(100)
        update_text.setPlainText("""【v2.0 更新说明】
• 新增术语自动翻译功能
• 支持免费翻译引擎（无需API密钥）
• 支持DeepSeek API翻译（高质量翻译）
• 支持批量术语翻译缓存机制""")
        update_info_layout.addWidget(update_text)
        layout.addWidget(update_info_group)
        
        # Cookie配置说明
        cookie_info_group = QGroupBox("Cookie配置（下载功能必需）")
        cookie_info_layout = QVBoxLayout(cookie_info_group)
        info_text2 = QTextEdit()
        info_text2.setReadOnly(True)
        info_text2.setMaximumHeight(120)
        info_text2.setPlainText("""⚠️ 下载功能需要配置微信公众平台的Cookies才能正常工作。

获取方法：
1. 使用Chrome浏览器登录 https://mp.weixin.qq.com
2. 按 F12 → Application → Cookies → https://mp.weixin.qq.com
3. 找到 'wxuin' 和 'pass_ticket' 的值，复制到下方。""")
        cookie_info_layout.addWidget(info_text2)
        layout.addWidget(cookie_info_group)
        
        cookie_group = QGroupBox("Cookies配置")
        cookie_layout = QGridLayout(cookie_group)
        cookie_layout.addWidget(QLabel("wxuin:"), 0, 0)
        self.wxuin_input = QLineEdit()
        cookie_layout.addWidget(self.wxuin_input, 0, 1, 1, 3)
        cookie_layout.addWidget(QLabel("pass_ticket:"), 1, 0)
        self.pass_ticket_input = QLineEdit()
        cookie_layout.addWidget(self.pass_ticket_input, 1, 1, 1, 3)
        cookie_layout.addWidget(QLabel("cookies文件路径:"), 2, 0)
        self.cookie_file_input = QLineEdit()
        cookie_layout.addWidget(self.cookie_file_input, 2, 1)
        self.browse_cookie_btn = QPushButton("浏览")
        self.browse_cookie_btn.clicked.connect(self.browse_cookie_file)
        cookie_layout.addWidget(self.browse_cookie_btn, 2, 2)
        self.load_cookie_btn = QPushButton("加载Cookies")
        self.load_cookie_btn.clicked.connect(self.load_cookies)
        cookie_layout.addWidget(self.load_cookie_btn, 2, 3)
        layout.addWidget(cookie_group)
        
        help_group = QGroupBox("📖 使用帮助")
        help_layout = QVBoxLayout(help_group)
        help_text = QTextEdit()
        help_text.setReadOnly(True)
        help_text.setMaximumHeight(150)
        help_text.setPlainText("""【术语表生成】
- 顺序提取模式：按文章顺序提取中英文对照词对
- 频率统计模式：统计英文单词出现频率，适合普通英文文章
- 术语翻译：提取后点击"识别并翻译术语"按钮自动补全中文翻译

【文章下载】
- 需要配置微信公众平台Cookies（wxuin和pass_ticket）
- 支持批量下载，自动保存为Markdown文件""")
        help_layout.addWidget(help_text)
        layout.addWidget(help_group)
        
        footer = QLabel("© 2026 June | Lingoes Family | Vibe Coding")
        footer.setAlignment(Qt.AlignCenter)
        footer.setStyleSheet("color: #888; font-size: 12px; padding: 10px;")
        layout.addWidget(footer)
        layout.addStretch()
        return tab

    def test_deepseek_api(self):
        """测试 DeepSeek API 连接"""
        api_key = self.deepseek_api_key_input.text().strip()
        if not api_key:
            QMessageBox.warning(self, "提示", "请输入 API Key 后再测试")
            return
        
        translator = DeepSeekTranslator(api_key)
        result = translator.translate_single("Hello")
        if result and result != "":
            QMessageBox.information(self, "测试成功", f"API 连接正常！\n测试翻译：Hello → {result}")
        else:
            QMessageBox.warning(self, "测试失败", "API 连接失败，请检查 API Key 是否正确")

    def setup_signals(self):
        self.downloader.progress_signal.connect(self.update_progress)
        self.downloader.log_signal.connect(self.log)
        self.downloader.article_downloaded.connect(self.on_article_downloaded)
        self.enable_translate_check.toggled.connect(self.on_translate_toggled)
        self.translate_engine.currentIndexChanged.connect(self.on_translate_engine_changed)
        self.format_combo_order.currentIndexChanged.connect(self.on_format_changed)

    def on_translate_toggled(self, checked):
        """启用/关闭自动翻译"""
        self.term_processor.enable_auto_translate = checked
        if checked:
            self.log("术语自动翻译已启用")
            # 更新翻译引擎设置
            self.on_translate_engine_changed(self.translate_engine.currentIndex())
        else:
            self.log("术语自动翻译已关闭")

    def on_translate_engine_changed(self, index):
        """切换翻译引擎"""
        if index == 0:  # 免费引擎
            self.term_processor.translator_mode = "free"
            self.log("翻译引擎已切换到：免费翻译引擎")
        else:  # DeepSeek API
            api_key = self.deepseek_api_key_input.text().strip()
            if api_key:
                self.term_processor.translator_mode = "deepseek"
                self.term_processor.set_deepseek_key(api_key)
                self.log("翻译引擎已切换到：DeepSeek API")
            else:
                self.log("警告：DeepSeek API Key 未设置，将使用免费翻译引擎")
                self.term_processor.translator_mode = "free"

    def on_format_changed(self, index):
        """顺序提取模式格式改变时更新处理器"""
        self.term_processor.format_type = index

    def browse_save_path(self):
        path = QFileDialog.getExistingDirectory(self, "选择保存目录")
        if path:
            self.save_path.setText(path)

    def browse_cookie_file(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "选择Cookie文件", "", "JSON文件 (*.json);;所有文件 (*)")
        if file_path:
            self.cookie_file_input.setText(file_path)

    def load_cookies(self):
        cookies = {}
        if self.wxuin_input.text():
            cookies['wxuin'] = self.wxuin_input.text()
        if self.pass_ticket_input.text():
            cookies['pass_ticket'] = self.pass_ticket_input.text()
        if cookies:
            self.downloader.cookies = cookies
            self.log("Cookies已加载")
        elif self.cookie_file_input.text():
            self.downloader.init_api(self.cookie_file_input.text())
        else:
            self.log("请至少提供wxuin和pass_ticket，或指定cookies文件")

    def start_download(self):
        urls_text = self.url_input.toPlainText().strip()
        if not urls_text:
            self.log("请输入文章URL")
            return
        urls = [url.strip() for url in urls_text.split('\n') if url.strip()]
        if not urls:
            return
        cookies = {}
        if self.wxuin_input.text():
            cookies['wxuin'] = self.wxuin_input.text()
        if self.pass_ticket_input.text():
            cookies['pass_ticket'] = self.pass_ticket_input.text()
        if cookies:
            self.downloader.cookies = cookies
        save_dir = self.save_path.text() or "./downloads"
        self.download_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress_bar.setRange(0, len(urls))
        self.progress_bar.setValue(0)
        self.download_results = []
        self._download_thread = DownloadWorker(self.downloader, urls, save_dir, self.max_workers.value())
        self._download_thread.finished.connect(self.on_download_finished)
        self._download_thread.start()

    def on_download_finished(self):
        self.download_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status_label.setText("下载完成")

    def stop_download(self):
        self.downloader.stop()
        self.log("用户请求停止下载")

    def update_progress(self, current, total):
        self.progress_bar.setValue(current)
        self.status_label.setText(f"正在下载... {current}/{total}")
        if current == total:
            self.status_label.setText("下载完成")

    def on_article_downloaded(self, article_info):
        self.download_results.append(article_info)
        if article_info['success'] and article_info['path']:
            item_text = f"📄 {article_info['title']} | {article_info['path']}"
            self.articles_list.addItem(item_text)
            self.articles_list.item(self.articles_list.count() - 1).setData(Qt.UserRole, article_info['path'])

    def log(self, message):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {message}")
        self.status_label.setText(message)

    def import_downloaded_to_term(self):
        save_dir = self.save_path.text()
        if not os.path.exists(save_dir):
            self.log("下载目录不存在，请先下载文章")
            return
        self.articles_list.clear()
        count = 0
        for root, dirs, files in os.walk(save_dir):
            for file in files:
                if file.endswith('.md'):
                    filepath = os.path.join(root, file)
                    title = os.path.splitext(file)[0]
                    item = QListWidgetItem(f"📄 {title} | {filepath}")
                    item.setData(Qt.UserRole, filepath)
                    self.articles_list.addItem(item)
                    count += 1
        self.log(f"已加载 {count} 篇已下载文章")
        if count > 0:
            self.process_selected_btn.setEnabled(True)

    def load_downloaded_articles(self):
        self.import_downloaded_to_term()

    def process_selected_articles(self):
        selected = self.articles_list.selectedItems()
        if not selected:
            self.log("请先选择要处理的文章")
            return
        selected.sort(key=lambda item: self.articles_list.row(item))
        all_pairs = []
        file_paths = []
        for item in selected:
            filepath = item.data(Qt.UserRole)
            if not filepath or not os.path.exists(filepath):
                continue
            file_paths.append(filepath)
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    content = f.read()
                clean = re.sub(r'^#+\s+.*$', '', content, flags=re.MULTILINE)
                clean = re.sub(r'^\s*[`*\-_~]+.*$', '', clean, flags=re.MULTILINE)
                pairs = self.term_processor.extract_ordered_pairs(clean)
                all_pairs.extend(pairs)
                self.log(f"从 {os.path.basename(filepath)} 提取到 {len(pairs)} 个词对")
            except Exception as e:
                self.log(f"读取失败 {filepath}: {e}")
        if not all_pairs:
            self.log("未提取到任何中英文词对")
            self.result_preview.setText("未提取到任何中英文词对")
            return
        self.current_terms_ordered = all_pairs
        self.current_is_ordered = True
        preview = self.term_processor.generate_markdown_table(all_pairs, title="顺序术语表", is_ordered=True)
        self.result_preview.setText(preview)
        self.log(f"成功提取 {len(all_pairs)} 个词对")
        summary = f"[已合并 {len(file_paths)} 篇文章，共提取 {len(all_pairs)} 个词对]\n"
        for idx, (en, cn) in enumerate(all_pairs[:20], 1):
            summary += f"{idx}. {en}  {cn}\n"
        if len(all_pairs) > 20:
            summary += "..."
        self.term_input.setText(summary)
        self.tabs.setCurrentIndex(1)

    def process_single_article(self, item):
        filepath = item.data(Qt.UserRole)
        if not filepath or not os.path.exists(filepath):
            self.log(f"文件不存在: {filepath}")
            return
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                content = f.read()
            clean = re.sub(r'^#+\s+.*$', '', content, flags=re.MULTILINE)
            clean = re.sub(r'^\s*[`*\-_~]+.*$', '', clean, flags=re.MULTILINE)
            self.term_input.setText(clean)
            self.process_current_text()
            self.log(f"已处理: {os.path.basename(filepath)}")
        except Exception as e:
            self.log(f"处理失败: {str(e)}")

    def process_current_text(self):
        text = self.term_input.toPlainText()
        if not text.strip():
            self.log("请输入文章内容")
            return
        if self.mode_ordered.isChecked():
            pairs = self.term_processor.extract_ordered_pairs(text)
            if not pairs:
                self.log("未提取到中英文词对")
                self.result_preview.setText("未提取到任何中英文词对")
                return
            self.current_terms_ordered = pairs
            self.current_is_ordered = True
            preview = self.term_processor.generate_markdown_table(pairs, title="顺序术语表", is_ordered=True)
            self.result_preview.setText(preview)
            self.log(f"提取到 {len(pairs)} 个词对")
        else:
            self.term_processor.min_word_len = self.min_len.value()
            self.term_processor.min_freq = self.min_freq.value()
            self.term_processor.max_terms = self.max_terms.value()
            stopwords_text = self.stopwords_input.text().strip()
            if stopwords_text:
                custom = [w.strip().lower() for w in stopwords_text.split(',') if w.strip()]
                self.term_processor.stopwords = self.term_processor.DEFAULT_STOPWORDS.copy()
                self.term_processor.stopwords.update(custom)
            else:
                self.term_processor.stopwords = self.term_processor.DEFAULT_STOPWORDS.copy()
            method = "nlp" if self.extract_mode.currentIndex() == 1 else "basic"
            if method == "nlp":
                terms = self.term_processor.extract_nlp(text)
            else:
                terms = self.term_processor.extract_basic(text)
            if not terms:
                self.log("未提取到术语，请调整参数或检查文章是否包含英文")
                self.result_preview.setText("未提取到符合条件的术语")
                return
            self.current_terms_freq = terms
            self.current_is_ordered = False
            preview = self.term_processor.generate_markdown_table(terms, title="频率术语表", is_ordered=False)
            self.result_preview.setText(preview)
            self.log(f"提取到 {len(terms)} 个术语")

    def translate_current_terms(self):
        """手动触发当前术语的翻译"""
        if self.mode_ordered.isChecked():
            if not hasattr(self, 'current_terms_ordered') or not self.current_terms_ordered:
                self.log("请先提取术语（点击【处理选中文章】或【提取术语】）")
                return
            terms = self.current_terms_ordered
            is_ordered = True
        else:
            if not hasattr(self, 'current_terms_freq') or not self.current_terms_freq:
                self.log("请先提取术语")
                return
            terms = self.current_terms_freq
            is_ordered = False
        
        # 启用翻译功能
        self.term_processor.enable_auto_translate = True
        self.enable_translate_check.setChecked(True)
        
        # 更新翻译引擎设置
        if self.translate_engine.currentIndex() == 1:
            api_key = self.deepseek_api_key_input.text().strip()
            if api_key:
                self.term_processor.translator_mode = "deepseek"
                self.term_processor.set_deepseek_key(api_key)
            else:
                self.term_processor.translator_mode = "free"
        else:
            self.term_processor.translator_mode = "free"
        
        # 执行翻译
        self.log("正在翻译术语，请稍候...")
        QApplication.processEvents()
        
        if is_ordered:
            # 顺序模式：为每个术语翻译
            translated_terms = []
            for en, cn in terms:
                if not cn or cn.strip() == "":
                    translation = self.term_processor._translate_term(en)
                    translated_terms.append((en, translation if translation else ""))
                else:
                    translated_terms.append((en, cn))
            self.current_terms_ordered = translated_terms
            preview = self.term_processor.generate_markdown_table(translated_terms, title="顺序术语表（已翻译）", is_ordered=True)
        else:
            # 频率模式：批量翻译
            translated_terms = self.term_processor.translate_terms(terms)
            self.current_terms_freq = translated_terms
            preview = self.term_processor.generate_markdown_table(translated_terms, title="频率术语表（已翻译）", is_ordered=False)
        
        self.result_preview.setText(preview)
        self.log(f"翻译完成，共处理 {len(translated_terms)} 个术语")

    def ai_extract_terms(self):
        """AI术语提取：通过DeepSeek API自动理解并归纳文章术语"""
        # 获取API Key
        api_key = self.deepseek_api_key_input.text().strip()
        if not api_key:
            self.log("错误：请先在翻译设置中输入DeepSeek API Key")
            QMessageBox.warning(self, "缺少API Key", "请先在【术语翻译设置】中输入DeepSeek API Key")
            return
        
        # 获取文章内容
        text = self.term_input.toPlainText().strip()
        if not text:
            self.log("错误：请先输入文章内容或链接")
            QMessageBox.warning(self, "缺少内容", "请先在文章输入区输入文章内容")
            return
        
        # 获取输出格式偏好
        format_idx = self.format_combo_order.currentIndex()
        format_hint = "英文 中文" if format_idx == 0 else "中文 英文"
        
        self.log("正在调用AI提取术语，请稍候...")
        self.ai_extract_btn.setEnabled(False)
        QApplication.processEvents()
        
        try:
            # 构建AI提示词
            prompt = f"""你是一个专业的术语提取和翻译助手。请从以下文章中提取重要术语，并提供中文翻译。

要求：
1. 提取文章中的专业术语、关键词、重要概念
2. 每个术语提供英文和中文对照
3. 按照文章出现的顺序排列
4. 只提取有意义的术语，忽略普通词汇
5. 输出格式为：{format_hint}

文章内容：
{text[:8000]}

请输出术语列表，每行一个术语，格式为：英文术语 | 中文翻译"""

            # 调用DeepSeek API
            import requests as req
            response = req.post(
                "https://api.deepseek.com/v1/chat/completions",
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {api_key}"
                },
                json={
                    "model": "deepseek-chat",
                    "messages": [
                        {"role": "system", "content": "你是一个专业的术语提取和翻译助手，擅长从文章中提取专业术语并提供准确的中文翻译。"},
                        {"role": "user", "content": prompt}
                    ],
                    "temperature": 0.3,
                    "max_tokens": 4000
                },
                timeout=60
            )
            
            if response.status_code == 200:
                result = response.json()
                ai_response = result["choices"][0]["message"]["content"].strip()
                
                # 解析AI返回的术语列表
                terms = self._parse_ai_terms(ai_response, format_idx)
                
                if not terms:
                    self.log("AI未能提取到术语，请检查文章内容")
                    self.result_preview.setText("AI未能提取到术语，请检查文章内容或调整提示词")
                    return
                
                # 存储结果
                self.current_terms_ordered = terms
                self.current_is_ordered = True
                
                # 生成预览
                preview = self.term_processor.generate_markdown_table(terms, title="AI提取术语表", is_ordered=True)
                self.result_preview.setText(preview)
                self.log(f"AI提取完成，共 {len(terms)} 个术语")
            else:
                error_msg = response.json().get("error", {}).get("message", "未知错误")
                self.log(f"AI提取失败: {error_msg}")
                self.result_preview.setText(f"AI提取失败: {error_msg}")
                
        except Exception as e:
            self.log(f"AI提取异常: {str(e)}")
            self.result_preview.setText(f"AI提取异常: {str(e)}")
        finally:
            self.ai_extract_btn.setEnabled(True)
    
    def _parse_ai_terms(self, ai_response: str, format_idx: int) -> List[Tuple[str, str]]:
        """解析AI返回的术语文本"""
        terms = []
        lines = ai_response.split('\n')
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # 尝试多种分隔符
            for sep in ['|', '\t', ' - ', ' — ']:
                if sep in line:
                    parts = line.split(sep)
                    if len(parts) >= 2:
                        en = parts[0].strip()
                        cn = parts[1].strip()
                        if en and cn:
                            terms.append((en, cn))
                            break
            
            # 如果没有分隔符，尝试按空格分割（最后一个词作为中文）
            if not any(sep in line for sep in ['|', '\t', ' - ', ' — ']):
                # 尝试匹配 "英文 中文" 格式
                match = re.match(r'^([A-Za-z][A-Za-z\s\'\-/]+?)\s+([\u4e00-\u9fff]+)$', line)
                if match:
                    terms.append((match.group(1).strip(), match.group(2).strip()))
        
        return terms

    def save_terms(self):
        if self.mode_ordered.isChecked():
            if not hasattr(self, 'current_terms_ordered') or not self.current_terms_ordered:
                self.log("请先提取术语（点击【处理选中文章】或【提取术语】）")
                return
            terms = self.current_terms_ordered
            is_ordered = True
        else:
            if not hasattr(self, 'current_terms_freq') or not self.current_terms_freq:
                self.log("请先提取术语")
                return
            terms = self.current_terms_freq
            is_ordered = False
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "保存术语表", "",
            "Excel 文件 (*.xlsx);;Markdown 文件 (*.md);;JSON 文件 (*.json);;CSV 文件 (*.csv)"
        )
        if not file_path:
            return
        ext = os.path.splitext(file_path)[1].lower()
        fmt = self.output_format.currentIndex()
        try:
            if ext == '.xlsx' or fmt == 3:
                if not EXCEL_AVAILABLE:
                    QMessageBox.warning(self, "缺少依赖", "请安装 openpyxl")
                    return
                self.term_processor.generate_excel(terms, file_path, "术语表", is_ordered)
            elif ext == '.json' or fmt == 1:
                content = self.term_processor.generate_json(terms, is_ordered=is_ordered)
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)
            elif ext == '.csv' or fmt == 2:
                content = self.term_processor.generate_csv(terms, is_ordered)
                with open(file_path, 'w', encoding='utf-8-sig') as f:
                    f.write(content)
            else:
                content = self.term_processor.generate_markdown_table(terms, is_ordered=is_ordered)
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(content)
            self.log(f"保存成功: {file_path}")
        except Exception as e:
            self.log(f"保存失败: {e}")
            QMessageBox.critical(self, "错误", str(e))


def main():
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()